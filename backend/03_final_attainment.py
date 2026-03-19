from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import re

# ==============================
# VS CODE SETUP - DEFINE LOCAL FILE PATHS HERE
# Input files go in data/   |   Outputs go in outputs/
# ==============================
CO_FILE = os.getenv("CO_FILE", os.path.join("outputs", "CO_ATTAINMENT_FINAL.xlsx"))
TERMINAL_FILE = os.getenv("TERMINAL_FILE", os.path.join("data", "TERMINAL.xlsx"))
OUT_FILE = os.getenv("FINAL_OUT_FILE", os.path.join("outputs", "CO_ATTAINMENT_SHEET2_FINAL_ABSOLUTE_COMPLETE.xlsx"))

print("Files assigned:")
print("   CO File        ->", CO_FILE)
print("   Terminal File  ->", TERMINAL_FILE)

# ==============================
# LOAD WORKBOOKS
# ==============================
wb = load_workbook(CO_FILE, data_only=False)
sheet1 = wb["Sheet1"]
sheet2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.create_sheet("Sheet2")

term_wb = load_workbook(TERMINAL_FILE, data_only=False)
term_ws = term_wb.active


def env_float(name):
    raw = os.getenv(name)
    if raw is None or str(raw).strip() == "":
        return None
    return float(raw)


def get_ela_values():
    csv_values = os.getenv("ELA_VALUES")
    if csv_values:
        parts = [p.strip() for p in csv_values.split(",") if p.strip()]
        if len(parts) != 6:
            raise ValueError("ELA_VALUES must contain 6 comma-separated numbers for CO1..CO6")
        return [float(p) for p in parts]

    values = []
    for i in range(6):
        env_val = env_float(f"ELA_CO{i+1}")
        if env_val is not None:
            values.append(env_val)
            continue
        values.append(float(input(f"Enter ELA for CO{i+1}: ")))
    return values

# ==============================
# HELPERS
# ==============================
def rewrite_formula(formula):
    if not formula or not formula.startswith("="): return formula
    def repl(m): return f"Sheet1!{m.group(0)}"
    return "=" + re.sub(r'(\$?[A-Z]{1,3}\$?\d+)', repl, formula[1:])

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

START_ROW = 5
LAST_ROW = sheet1.max_row

# ==============================
# COPY A:C (IDENTITY)
# ==============================
for r in range(START_ROW, LAST_ROW + 1):
    if not sheet1.cell(r, 1).value: continue
    sheet2.cell(r, 1).value = sheet1.cell(r, 1).value
    sheet2.cell(r, 2).value = sheet1.cell(r, 2).value
    sheet2.cell(r, 3).value = sheet1.cell(r, 3).value
    sheet2.cell(r, 1).alignment = center
    sheet2.cell(r, 2).alignment = center
    sheet2.cell(r, 3).alignment = left

# ==============================
# INTERNAL %
# ==============================
sheet2.merge_cells("E3:J3")
sheet2["E3"] = "CO BASED Percentage of marks (Internal Assessment only)"
sheet2["E3"].alignment = center

SRC_START = 80  # CB
DST_START = 5   # E

for r in range(START_ROW, LAST_ROW + 1):
    if not sheet1.cell(r, SRC_START).value: continue
    for i in range(6):
        src = sheet1.cell(r, SRC_START + i)
        dst = sheet2.cell(r, DST_START + i)
        dst.value = rewrite_formula(src.value) if src.data_type == "f" else src.value
        dst.number_format = "0.00"
        dst.alignment = center

# ==============================
# TERMINAL MARKS
# ==============================
sheet2.merge_cells("L3:Q3")
sheet2["L3"] = "Terminal Assessment"
sheet2["L3"].alignment = center

for i in range(6):
    sheet2.cell(5, 12 + i).value = term_ws.cell(2, 7 + i).value
    sheet2.cell(5, 12 + i).alignment = center

for r in range(3, term_ws.max_row + 1):
    for i in range(6):
        cell = sheet2.cell(r + 4, 12 + i)
        cell.value = term_ws.cell(r, 7 + i).value
        cell.number_format = "0.00"
        cell.alignment = center

# ==============================
# FINAL CO (0.7 + 0.3)
# ==============================
for i in range(6):
    sheet2.cell(5, 19 + i).value = f"CO{i+1}"
    sheet2.cell(5, 19 + i).alignment = center

for r in range(8, LAST_ROW + 1):
    if not sheet2.cell(r, 2).value: continue
    for i in range(6):
        e = get_column_letter(5 + i)
        l = get_column_letter(12 + i)
        s = 19 + i
        cell = sheet2.cell(r, s)
        cell.value = f"=0.7*{e}{r}+0.3*{l}{r}"
        cell.number_format = "0.00"
        cell.alignment = center

# ==============================
# ATTAINMENT SUMMARY
# ==============================
last_student = 8
while sheet2.cell(last_student, 2).value: last_student += 1
last_student -= 1
summary_row = last_student + 3

sheet2.cell(summary_row, 18).value = "Total"
for i in range(6):
    col = get_column_letter(19 + i)
    sheet2.cell(summary_row, 19 + i).value = f"=COUNT({col}8:{col}{last_student})"

ep = env_float("EP_VALUE")
if ep is None:
    ep = float(input("\nEnter EP value: "))
sheet2.cell(summary_row + 1, 18).value = "EP"
for i in range(6):
    sheet2.cell(summary_row + 1, 19 + i).value = ep

constraint = os.getenv("CONSTRAINT_VALUE")
if constraint is None or str(constraint).strip() == "":
    constraint = input("Enter Constraint value (e.g. 79.99): ")
sheet2.cell(summary_row + 2, 18).value = "Constraint"
for i in range(6):
    col = get_column_letter(19 + i)
    sheet2.cell(summary_row + 2, 19 + i).value = f'=COUNTIF({col}8:{col}{last_student},">={constraint}")'

sheet2.cell(summary_row + 3, 18).value = "Actual Attainment (%)"
for i in range(6):
    c = get_column_letter(19 + i) + str(summary_row + 2)
    e = get_column_letter(19 + i) + str(summary_row + 1)
    cell = sheet2.cell(summary_row + 3, 19 + i)
    cell.value = f"=({c}/{e})*100"
    cell.number_format = "0.00"

# ==============================
# ELA & RELATIVE ATTAINMENT
# ==============================
ela_row = summary_row + 5
rel_row = ela_row + 1

sheet2.cell(ela_row, 18).value = "ELA"
sheet2.cell(rel_row, 18).value = "Relative Attainment (%)"

ela_values = get_ela_values()
for i in range(6):
    ela = ela_values[i]
    sheet2.cell(ela_row, 19 + i).value = ela

    actual = get_column_letter(19 + i) + str(summary_row + 3)
    ela_c = get_column_letter(19 + i) + str(ela_row)

    cell = sheet2.cell(rel_row, 19 + i)
    cell.value = f"=MIN(({ela_c}/{actual})*100,100)"
    cell.number_format = "0.00"
    cell.alignment = center

# ==============================
# SAVE
# ==============================
wb.save(OUT_FILE)
print(f"\nCO ATTAINMENT CLOSED - Relative Attainment calculated. Saved to {OUT_FILE}")