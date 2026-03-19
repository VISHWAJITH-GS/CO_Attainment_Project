from openpyxl import load_workbook
import os
import re

# ==============================
# VS CODE SETUP - DEFINE LOCAL FILE PATHS HERE
# Input files go in data/   |   Outputs go in outputs/
# ==============================
# TEMPLATE_FILE: name must match the _FINAL.xlsx produced by Script 1
TEMPLATE_FILE = os.getenv("TEMPLATE_FILE", os.path.join("data", "CO ATTAINMENT TEMPLATE (1).xlsx"))
CAT1_FILE = os.getenv("CAT1_FILE", os.path.join("data", "Cross platform -CAT 1 (1).xlsx"))
CAT2_FILE = os.getenv("CAT2_FILE", os.path.join("data", "Cross platform-CAT 2.xlsx"))
ASS1_FILE = os.getenv("ASS1_FILE", os.path.join("data", "Cross platform -Ass 1 (1).xlsx"))
ASS2_FILE = os.getenv("ASS2_FILE", os.path.join("data", "Cross platform -Ass 2 (1).xlsx"))
OUT_FILE = os.getenv("AGG_OUT_FILE", os.path.join("outputs", "CO_ATTAINMENT_FINAL.xlsx"))

template_wb = None
template_ws = None

# ==============================
# HELPERS
# ==============================
def normalize(q):
    if q is None: return None
    return re.sub(r'\.', '', str(q)).strip().upper()

def clean_numeric(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    if isinstance(val, str):
        s = val.strip().replace("'", "")
        if re.fullmatch(r"-?\d+", s): return int(s)
        if re.fullmatch(r"-?\d*\.\d+", s): return float(s)
    return None

# ==============================
# PROCESS CAT
# ==============================
def process_cat(cat_file, template_start_col, template_end_col):
    print(f"Processing CAT file: {cat_file}...")
    cat_wb = load_workbook(cat_file, data_only=True)
    cat_ws = cat_wb.active

    row = 13
    while cat_ws.cell(row=row, column=2).value: row += 1
    student_count = row - 13

    for i in range(student_count):
        template_ws.cell(row=8+i, column=2).value = cat_ws.cell(row=13+i, column=2).value
        template_ws.cell(row=8+i, column=3).value = cat_ws.cell(row=13+i, column=3).value

    cat_q, cat_max, cat_co = {}, {}, {}
    for col in range(4, cat_ws.max_column + 1):
        q = normalize(cat_ws.cell(row=10, column=col).value)
        if q:
            cat_q[q] = col
            cat_max[q] = clean_numeric(cat_ws.cell(row=11, column=col).value)
            cat_co[q]  = cat_ws.cell(row=12, column=col).value

    template_q = {}
    for col in range(template_start_col, template_end_col + 1):
        q = normalize(template_ws.cell(row=6, column=col).value)
        if q and q != "TOTAL":
            template_q[q] = col

    for q, tcol in template_q.items():
        if q in cat_q:
            template_ws.cell(row=5, column=tcol).value = cat_co[q]
            template_ws.cell(row=7, column=tcol).value = cat_max[q]
        else:
            template_ws.cell(row=5, column=tcol).value = None
            template_ws.cell(row=7, column=tcol).value = None

    for q, cat_col in cat_q.items():
        if q not in template_q: continue
        tcol = template_q[q]
        for i in range(student_count):
            template_ws.cell(row=8+i, column=tcol).value = clean_numeric(cat_ws.cell(row=13+i, column=cat_col).value)

# ==============================
# PROCESS ASSIGNMENT
# ==============================
def process_assignment(ass_file, source_start_col, template_start_col, max_cols=6, target_total=40):
    print(f"Processing Assignment file: {ass_file}...")
    ass_wb = load_workbook(ass_file, data_only=True)
    ass_ws = ass_wb.active

    row = 4
    while ass_ws.cell(row=row, column=source_start_col).value: row += 1
    student_count = row - 4

    ass_cols = ass_ws.max_column - source_start_col + 1
    if ass_cols > max_cols: raise ValueError(f"{ass_file} has more than {max_cols} columns")

    col_max_values = []
    for c in range(ass_cols):
        values = []
        for i in range(student_count):
            val = clean_numeric(ass_ws.cell(row=4+i, column=source_start_col+c).value)
            template_ws.cell(row=8+i, column=template_start_col+c).value = val
            if isinstance(val, (int, float)): values.append(val)
        col_max = max(values) if values else 0
        col_max_values.append(col_max)

    # --- Enforce total = 40 ---
    current_sum = sum(col_max_values)
    diff = target_total - current_sum
    if diff != 0 and col_max_values:
        col_max_values[-1] += diff  # adjust last column

    # Write max marks to template row 7
    for idx, val in enumerate(col_max_values):
        template_ws.cell(row=7, column=template_start_col+idx).value = val

def run_pipeline() -> str:
    global template_wb, template_ws

    print(f"Loading Template: {TEMPLATE_FILE}...")
    template_wb = load_workbook(TEMPLATE_FILE)
    template_ws = template_wb.worksheets[1]

    process_cat(CAT1_FILE, 5, 25)
    process_cat(CAT2_FILE, 27, 46)
    process_assignment(ASS1_FILE, 7, 56)
    process_assignment(ASS2_FILE, 7, 64)

    template_wb.save(OUT_FILE)
    print(f"Template populated successfully -> {OUT_FILE}")
    return OUT_FILE


if __name__ == "__main__":
    run_pipeline()